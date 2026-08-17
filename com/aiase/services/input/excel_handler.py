"""excel处理"""
import os

from com.aiase.services.input.base_handler import BaseHandler


class ExcelHandler(BaseHandler):
    """处理 .xlsx 工作表，逐sheet提取为表格文本"""

    def parse(self, file_path):
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts.append(f'## Sheet: {sheet}')
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else '' for c in row]
                if any(cells):
                    parts.append(' | '.join(cells))
        return '\n'.join(parts)
