"""Excel导出"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from com.aiase.entity.template import resolve_fields
from com.aiase.services.export.base_exporter import BaseExporter


class ExcelExporter(BaseExporter):

    def export(self, cases, template, path):
        fields = resolve_fields(template)
        wb = Workbook()
        ws = wb.active
        ws.title = '测试用例'
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_font = Font(color='FFFFFF', bold=True)
        ws.append(fields)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for c in cases:
            ws.append([c.get(f, '') for f in fields])
        # 含换行的单元格开启自动换行，保证步骤/结果逐行可见、一一对应
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value and '\n' in str(cell.value):
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        # 自动列宽
        for col in ws.columns:
            width = max(len(str(cell.value or '')) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(width, 50)
        wb.save(path)
        return path
